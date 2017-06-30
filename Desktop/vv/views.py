from django.shortcuts import render
import openpyxl
import subprocess
from .models import File
from .models import analyze
from django.core.files.storage import FileSystemStorage
from .forms import UploadFileForm
from openpyxl.styles import PatternFill

def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST , request.FILES)
        myfile = request.FILES['file']
        fs = FileSystemStorage()
        filename = fs.save(myfile.name,myfile )
        files = File.objects.all()
        return render(request , 'upload.html',{'form':form,'files':files})
    else:
        form = UploadFileForm()
    return render(request, 'upload.html' , {'form':form})


def change(request):
    if request.method == "POST":
        workbk = request.POST['workbk']
        no_sub = int(request.POST['no_sub'])
        new_file= request.POST['new_file']
        credit = request.POST['credit']
        print subprocess.check_output("pwd",shell=True)

        wb = openpyxl.load_workbook('/home/black/Desktop/mk/k/media/'+workbk)
        sheet = wb.active
        #max row
        mx_r = sheet.max_row
        #max Column
        mx_c = sheet.max_column
        #start and end range for sheet
        start_range = 'C2'
        end_range=chr(ord('C')+(no_sub-1))+str(mx_r)
        for row in sheet[start_range:end_range]:
            for col in row:
                if(col.value=='A'):
                    sheet[col.coordinate]=9
                elif(col.value=='B'):
                    sheet[col.coordinate]=8
                elif(col.value=='C'):
                    sheet[col.coordinate]=7
                elif(col.value=='D'):
                    sheet[col.coordinate]=6
                elif(col.value=='E'):
                    sheet[col.coordinate]=5
                elif(col.value=='S'):
                    sheet[col.coordinate]=10
                else:
                    sheet[col.coordinate]=10.1
		wb.save(filename="/home/black/Desktop/mk/k/media/step1.xlsx")

        uploaded = hitler(credit,no_sub,mx_r,mx_c,start_range,end_range)
        print subprocess.check_output("cp /home/black/Desktop/mk/k/media/step2.xlsx /home/black/Desktop/mk/k/analysis/templates" , shell=True)
        return render(request,'analyze.html',{'uploaded':uploaded})
    else:
        return render(request,'analyze.html')

        #workbk1 = '/home/darkman/Desktop/mk/k/media/step1.xlsx'
        #the result is stored in this location in sheet
def hitler(credit,no_sub,mx_r,mx_c,start_range,end_range):

    wb = openpyxl.load_workbook('/home/black/Desktop/mk/k/media/step1.xlsx')
    sheet = wb.active
    #
    gpa = chr(ord('C')+no_sub)
    print gpa
    credit = list(credit)
    print credit
    #flag
    k = 2
    if((mx_c-2)==no_sub):
        for i in sheet[start_range:end_range]:
            sum1 = 0
            crdt = 0
            cr_sum = 0
            for col in i:
                sum1 = sum1 + ((col.value)*int(credit[crdt]))
                cr_sum = cr_sum + int(credit[crdt])
                crdt = crdt  + 1
            print cr_sum
            val = gpa + str(k)
            k = k+1
            res = round(float(sum1)/cr_sum , 2)
            sheet[val]=res
            sheet[val].fill = PatternFill(bgColor="FF0900" )
            print (val , sheet[val],sheet[val].value)
            wb.save(filename="/home/black/Desktop/mk/k/media/step2.xlsx")

            filepath = 'media/step2.xlsx'
    return filepath

def update(request):
    if request.method == "POST":
        workbk = request.POST['workbk']
        reg = request.POST['reg']
        sub = request.POST['sub']
        grd  = request.POST['grd']
        wb = openpyxl.load_workbook('/home/black/Desktop/mk/k/media/'+workbk)
        sheet = wb.active
        mx_r = sheet.max_row
        mx_c = sheet.max_column
        end_reg = 'A'+str(mx_r)
        i = 0
        for row in sheet['A1':end_reg]:
            i = i+1
            for col in row:
                if (col.value == int(reg)):
                    r = i
                    break
        end_col  = chr((ord('A')+mx_c)-1)+'1'
        print end_col
        j=0
        for row in sheet['A1':end_col]:
            for col in row:
                j = j+1
                if(col.value == sub):
                    c = j


        print r,c
        print "the value is " + sheet.cell(row=r,column=c).value
        sheet.cell(row=r,column=c).value = grd
        print "the value is " + sheet.cell(row=r,column=c).value

        file_n = workbk + '-updated.xlsx'
        wb.save(filename=file_n)
        return render(request,'update.html')
    else:
        return render(request,'update.html')
